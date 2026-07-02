import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

wb = openpyxl.Workbook()

# ── 공통 스타일 ─────────────────────────────────────────────
header_font  = Font(bold=True, color='FFFFFF')
header_fill  = PatternFill('solid', fgColor='4472C4')
hint_fill    = PatternFill('solid', fgColor='D9E1F2')
hint_font    = Font(color='595959', italic=True)
lock_fill    = PatternFill('solid', fgColor='FFF2CC')
sample_fill  = PatternFill('solid', fgColor='E2EFDA')   # 연두: 샘플 데이터
notice_fill  = PatternFill('solid', fgColor='FCE4D6')   # 연주황: 안내 행
notice_font  = Font(color='C00000', bold=True)
center       = Alignment(horizontal='center', vertical='center')
left         = Alignment(horizontal='left',   vertical='center')

# ── Sheet2: 코드표 ──────────────────────────────────────────
ws2 = wb.active
ws2.title = '코드표'

ws2.append(['응답범위', '응답범위코드'])
ws2.append(['본인', '01'])
ws2.append(['상사', '02'])
ws2.append(['구성원', '03'])
ws2.append(['동료', '04'])

for cell in ['A1', 'B1']:
    ws2[cell].font = header_font
    ws2[cell].fill = header_fill
    ws2[cell].alignment = center

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 16
ws2.row_dimensions[1].height = 22

# ── Named Range: RESP_NAMES ──────────────────────────────────
dn = DefinedName('RESP_NAMES', attr_text="코드표!$A$2:$A$5")
wb.defined_names['RESP_NAMES'] = dn

# ── Sheet1: 업로드_템플릿 ────────────────────────────────────
ws1 = wb.create_sheet('업로드_템플릿', 0)

# ── 1행: 파일 사용 안내 (병합) ──────────────────────────────
ws1.merge_cells('A1:D1')
notice_cell = ws1['A1']
notice_cell.value = (
    '✅ 작성 방법: A·B열에 아이디를 입력하고 C열에서 응답범위를 드롭다운으로 선택하세요. '
    'D열(응답범위코드)은 자동 입력되므로 수정하지 마세요. '
    '3~5행은 샘플 데이터입니다. 실제 등록 시 삭제 후 사용하세요.'
)
notice_cell.fill    = notice_fill
notice_cell.font    = notice_font
notice_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws1.row_dimensions[1].height = 42

# ── 2행: 컬럼 헤더 ──────────────────────────────────────────
headers = ['대상자 아이디', '응답자 아이디', '응답범위', '응답범위코드']
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
ws1.row_dimensions[2].height = 22

# ── 3행: 입력 힌트 ──────────────────────────────────────────
hints = ['예: hong@hunet.co.kr', '예: kim@hunet.co.kr', '▼ 드롭다운 선택', '자동 입력 (수정 금지)']
for col_idx, hint in enumerate(hints, 1):
    cell = ws1.cell(row=3, column=col_idx, value=hint)
    cell.fill      = hint_fill
    cell.font      = hint_font
    cell.alignment = center
ws1.row_dimensions[3].height = 20

# ── 4~8행: 샘플 데이터 ──────────────────────────────────────
samples = [
    ('hong@hunet.co.kr',  'hong@hunet.co.kr',  '본인'),
    ('hong@hunet.co.kr',  'kim@hunet.co.kr',   '상사'),
    ('hong@hunet.co.kr',  'lee@hunet.co.kr',   '구성원'),
    ('hong@hunet.co.kr',  'park@hunet.co.kr',  '구성원'),
    ('hong@hunet.co.kr',  'choi@hunet.co.kr',  '동료'),
]
SAMPLE_START = 4
for i, (target, resp, role) in enumerate(samples):
    r = SAMPLE_START + i
    ws1.cell(row=r, column=1, value=target).fill = sample_fill
    ws1.cell(row=r, column=2, value=resp).fill   = sample_fill
    ws1.cell(row=r, column=3, value=role).fill   = sample_fill
    d_cell = ws1.cell(row=r, column=4,
                      value=f'=IF($C{r}="","",IFERROR(VLOOKUP($C{r},코드표!$A:$B,2,FALSE),""))')
    d_cell.fill = lock_fill
    for col in range(1, 5):
        ws1.cell(row=r, column=col).alignment = center

# ── 9~200행: 실제 입력 영역 ─────────────────────────────────
MAX_ROW = 200
for r in range(SAMPLE_START + len(samples), MAX_ROW + 1):
    d_cell = ws1.cell(row=r, column=4,
                      value=f'=IF($C{r}="","",IFERROR(VLOOKUP($C{r},코드표!$A:$B,2,FALSE),""))')
    d_cell.fill      = lock_fill
    d_cell.alignment = center

# ── 열 너비 / 2행 고정 ──────────────────────────────────────
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 20
ws1.freeze_panes = 'A4'  # 1~3행 고정

# ── C열 드롭다운 (응답범위) ──────────────────────────────────
dv = DataValidation(
    type='list',
    formula1='RESP_NAMES',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='응답범위 오류',
    error='코드표의 응답범위를 선택해주세요.'
)
dv.sqref = f'C4:C{MAX_ROW}'
ws1.add_data_validation(dv)

# ── 저장 ────────────────────────────────────────────────────
out_path = r'C:\Users\ke_1989\Documents\claudecode\labs-edms-prototype\진단\대상자_일괄등록_템플릿.xlsx'
wb.save(out_path)
print(f'저장 완료: {out_path}')
