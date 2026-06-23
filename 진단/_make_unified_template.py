"""
진단 일괄 등록 - 단일 시트 통합 템플릿 (1행 = 1질문).

질문유형 3종:
  척도형 / A/B 강제선택 / 서술형

핵심:
- 한 행 = 한 질문. 유형명/역량군명/세부역량명을 보고 시스템이 상위 계층 자동 생성·연결.
- 코드값/질문번호/글자수제한 모두 자동 처리 (사용자 입력 불필요).
- 척도형 → 척도단계 + 보기1~N 라벨 (기본 매우 그렇다 ~ 전혀 그렇지 않다, 수정 가능)
- A/B 강제선택 → A 좌측·B 우측 문구
- 서술형 → 질문내용만
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUTPUT = r"C:\Users\ke_1989\Documents\claudecode\labs-edms-prototype\진단\진단_일괄등록_템플릿.xlsx"

FONT = "맑은 고딕"

FILL_TITLE = PatternFill("solid", start_color="1F4E78")
FILL_GUIDE = PatternFill("solid", start_color="F2F2F2")
FILL_HEADER = PatternFill("solid", start_color="2E75B6")
FILL_SUBHEAD_PARENT = PatternFill("solid", start_color="DDEBF7")
FILL_SUBHEAD_Q = PatternFill("solid", start_color="DCEEDC")
FILL_SUBHEAD_SCALE_OPT = PatternFill("solid", start_color="EFE4F5")
FILL_SUBHEAD_AB = PatternFill("solid", start_color="FDE6E2")

# 질문유형별 행 색상 (3종)
FILL_QTYPE = {
    "척도형":       PatternFill("solid", start_color="EFE4F5"),
    "A/B 강제선택": PatternFill("solid", start_color="FDE6E2"),
    "서술형":       PatternFill("solid", start_color="F2F4F8"),
}

QTYPES = list(FILL_QTYPE.keys())

thin = Side(style="thin", color="C0C8D2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def font(size=10, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)


# 척도형 기본 라벨 (5점 기준)
LIKERT_5_LABELS = ['매우 그렇다', '그렇다', '보통이다', '그렇지 않다', '전혀 그렇지 않다']
LIKERT_5_SCORES = [5, 4, 3, 2, 1]


wb = Workbook()
ws = wb.active
ws.title = "진단 일괄 등록"
ws.sheet_view.showGridLines = False

# ── 컬럼 정의 ──
COLS = [
    # 소속 (자동 매핑 기준)
    ("★ 유형명",        20, "parent", "예: 본인 진단. 같은 이름이 다른 행에 또 나오면 같은 유형으로 묶임."),
    ("★ 역량군명",      16, "parent", "예: Self"),
    ("★ 세부역량명",    16, "parent", "예: 자기인식"),
    # 질문
    ("★ 질문유형",      14, "q",      "척도형 / A/B 강제선택 / 서술형"),
    ("★ 질문내용",      48, "q",      "사용자에게 노출되는 질문 텍스트"),
    # 척도형 전용: 보기 1~5
    ("보기1 내용",       18, "scale", "척도형 전용. 기본 '매우 그렇다' — 다른 값으로 수정 가능"),
    ("점수1",             6, "scale", "기본 5"),
    ("보기2 내용",       18, "scale", "기본 '그렇다'"),
    ("점수2",             6, "scale", "기본 4"),
    ("보기3 내용",       18, "scale", "기본 '보통이다'"),
    ("점수3",             6, "scale", "기본 3"),
    ("보기4 내용",       18, "scale", "기본 '그렇지 않다'"),
    ("점수4",             6, "scale", "기본 2"),
    ("보기5 내용",       18, "scale", "기본 '전혀 그렇지 않다'"),
    ("점수5",             6, "scale", "기본 1"),
    # A/B 강제선택 전용
    ("A 좌측 문구",      24, "ab",    "A/B 강제선택 전용. 양극의 한쪽 문구"),
    ("B 우측 문구",      24, "ab",    "A/B 강제선택 전용. 반대쪽 문구"),
    # 척도단계
    ("척도단계",          8, "scale", "척도형 전용. 2~10 (기본 5). 변경 시 보기 개수도 함께 조정."),
    # 비고
    ("비고",             20, "etc",   "내부 메모 (선택)"),
]

NCOL = len(COLS)

# 1행: 타이틀
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
c = ws.cell(row=1, column=1, value="진단 일괄 등록 템플릿  (한 줄 = 한 질문, 유형·역량군·세부역량은 자동 생성)")
c.font = font(14, True, "FFFFFF")
c.fill = FILL_TITLE
c.alignment = CENTER
ws.row_dimensions[1].height = 28

# 안내
guides = [
    "📌 작성 방법",
    "① 한 행에 한 질문을 적습니다. 그 행의 [유형명 / 역량군명 / 세부역량명]이 같으면 시스템이 같은 계층으로 묶어 유형·역량군·세부역량을 자동 생성합니다.",
    "② 질문유형 3종: ❶ 척도형(5점 기본, 보기 라벨 수정 가능) ❷ A/B 강제선택 ❸ 서술형",
    "③ 척도형은 '매우 그렇다 / 그렇다 / 보통이다 / 그렇지 않다 / 전혀 그렇지 않다'가 기본값으로 채워져 있고, 필요하면 셀을 수정해 다른 라벨로 바꿀 수 있어요. 척도단계도 조정 가능.",
    "④ A/B 강제선택은 A·B 두 문구만 작성. 서술형은 질문내용만 작성하면 됩니다.",
    "⑤ 행 색상은 질문유형 표시용 — 보라=척도형, 살구=A/B 강제선택, 회색=서술형",
]
for i, txt in enumerate(guides, start=2):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=NCOL)
    cell = ws.cell(row=i, column=1, value=txt)
    cell.font = font(10, bold=(i == 2), color="333333")
    cell.alignment = LEFT
    cell.fill = FILL_GUIDE
    ws.row_dimensions[i].height = 18 if i > 2 else 22

# 그룹 헤더 / 컬럼 헤더 행
GROUP_ROW = 8
HEADER_ROW = 9
ws.row_dimensions[7].height = 6

def merge_group(label, fill, start_col, end_col):
    ws.merge_cells(start_row=GROUP_ROW, start_column=start_col, end_row=GROUP_ROW, end_column=end_col)
    c = ws.cell(row=GROUP_ROW, column=start_col, value=label)
    c.font = font(10, True, "1F3A6E")
    c.fill = fill
    c.alignment = CENTER
    c.border = BORDER

merge_group("소속 (자동 매핑 기준)", FILL_SUBHEAD_PARENT, 1, 3)
merge_group("질문", FILL_SUBHEAD_Q, 4, 5)
merge_group("척도형 보기 1~5 (기본값 채워져 있고 수정 가능)", FILL_SUBHEAD_SCALE_OPT, 6, 15)
merge_group("A/B 강제선택 전용", FILL_SUBHEAD_AB, 16, 17)
merge_group("척도형 단계", FILL_SUBHEAD_SCALE_OPT, 18, 18)
merge_group("기타", FILL_GUIDE, 19, 19)
ws.row_dimensions[GROUP_ROW].height = 22

for i, (h, w, g, tip) in enumerate(COLS, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = font(10, True, "FFFFFF")
    c.fill = FILL_HEADER
    c.alignment = CENTER
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
    if tip:
        c.comment = Comment(tip, "guide")
ws.row_dimensions[HEADER_ROW].height = 32


# ── 샘플 데이터 ──
# 컬럼: 유형명, 역량군명, 세부역량명, 질문유형, 질문내용,
#       보기1내용, 점수1, ..., 보기5내용, 점수5,
#       A좌측문구, B우측문구, 척도단계, 비고
SAMPLES = [
    # 본인 진단 > Self > 자기인식 — 척도형(기본 라벨 그대로)
    ("본인 진단", "Self", "자기인식", "척도형",
     "나는 나의 강점과 약점을 명확히 인식한다.",
     LIKERT_5_LABELS[0], LIKERT_5_SCORES[0],
     LIKERT_5_LABELS[1], LIKERT_5_SCORES[1],
     LIKERT_5_LABELS[2], LIKERT_5_SCORES[2],
     LIKERT_5_LABELS[3], LIKERT_5_SCORES[3],
     LIKERT_5_LABELS[4], LIKERT_5_SCORES[4],
     "", "",
     5,
     "5점 척도 — 기본 보기 그대로"),
    # 척도형(보기 라벨 커스텀)
    ("본인 진단", "Self", "자기인식", "척도형",
     "자기관리 수준을 평가하면?",
     "매우 우수", 5, "우수", 4, "보통", 3, "부족", 2, "매우 부족", 1,
     "", "",
     5,
     "보기 라벨 수정 예시"),
    # 본인 진단 > Self > 자기인식 — A/B 강제선택
    ("본인 진단", "Self", "자기인식", "A/B 강제선택",
     "본인의 의사결정 방식에 더 가까운 쪽을 선택하세요.",
     "", "", "", "", "", "", "", "", "", "",
     "원칙을 우선시한다.", "상황에 맞게 유연하게 대응한다.",
     "",
     "A/B 강제선택"),
    # 본인 진단 > Self > 자기인식 — 서술형
    ("본인 진단", "Self", "자기인식", "서술형",
     "본인의 자기관리 방식에 대해 자유롭게 서술하세요.",
     "", "", "", "", "", "", "", "", "", "",
     "", "",
     "",
     "서술형"),
    # 본인 진단 > Self > 자기관리 — 척도형 2개
    ("본인 진단", "Self", "자기관리", "척도형",
     "나는 계획한 일을 끝까지 완수한다.",
     LIKERT_5_LABELS[0], LIKERT_5_SCORES[0],
     LIKERT_5_LABELS[1], LIKERT_5_SCORES[1],
     LIKERT_5_LABELS[2], LIKERT_5_SCORES[2],
     LIKERT_5_LABELS[3], LIKERT_5_SCORES[3],
     LIKERT_5_LABELS[4], LIKERT_5_SCORES[4],
     "", "",
     5,
     ""),
    # 구성원 진단 > Self > 자기인식
    ("구성원 진단", "Self", "자기인식", "척도형",
     "리더는 자신의 강점과 약점을 잘 알고 있다.",
     LIKERT_5_LABELS[0], LIKERT_5_SCORES[0],
     LIKERT_5_LABELS[1], LIKERT_5_SCORES[1],
     LIKERT_5_LABELS[2], LIKERT_5_SCORES[2],
     LIKERT_5_LABELS[3], LIKERT_5_SCORES[3],
     LIKERT_5_LABELS[4], LIKERT_5_SCORES[4],
     "", "",
     5,
     ""),
    ("구성원 진단", "Self", "자기인식", "서술형",
     "리더의 자기인식과 관련하여 인상 깊었던 일화가 있다면 적어주세요.",
     "", "", "", "", "", "", "", "", "", "",
     "", "",
     "",
     ""),
]

DATA_START = HEADER_ROW + 1
LEFT_COLS = {1, 2, 3, 5, 6, 8, 10, 12, 14, 16, 17, 19}

for ri, row in enumerate(SAMPLES, start=DATA_START):
    qtype = row[3]
    fill = FILL_QTYPE.get(qtype, None)
    for ci, v in enumerate(row, start=1):
        cell = ws.cell(row=ri, column=ci, value=v if v != "" else None)
        cell.font = font(10)
        cell.alignment = LEFT if ci in LEFT_COLS else CENTER
        cell.border = BORDER
        if fill:
            cell.fill = fill
    ws.row_dimensions[ri].height = 26

# 빈 행
SAMPLE_END = DATA_START + len(SAMPLES) - 1
EMPTY_END = SAMPLE_END + 60
for ri in range(SAMPLE_END + 1, EMPTY_END + 1):
    for ci in range(1, NCOL + 1):
        cell = ws.cell(row=ri, column=ci, value=None)
        cell.font = font(10)
        cell.alignment = LEFT if ci in LEFT_COLS else CENTER
        cell.border = BORDER

# ── 데이터 유효성 ──
dv_qtype = DataValidation(type="list", formula1=f'"{",".join(QTYPES)}"', allow_blank=True)
dv_qtype.add(f"D{DATA_START}:D{EMPTY_END}")
dv_qtype.error = "척도형 / A/B 강제선택 / 서술형 중에서 선택하세요"
dv_qtype.errorTitle = "잘못된 질문유형"
ws.add_data_validation(dv_qtype)

dv_scale = DataValidation(type="whole", operator="between", formula1=2, formula2=10, allow_blank=True)
dv_scale.add(f"R{DATA_START}:R{EMPTY_END}")
ws.add_data_validation(dv_scale)

ws.freeze_panes = f"D{DATA_START}"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Header row: {HEADER_ROW}, Data start: {DATA_START}, Sample rows: {len(SAMPLES)}, Total cols: {NCOL}")
